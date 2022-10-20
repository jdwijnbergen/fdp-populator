import Config
import openpyxl
from resource_classes import VPOrganisation, VPBiobank, VPPatientregistry, VPDataset, VPDistribution

class VPTemplateReader:
    """
    NOTE: this class is based on the folling specification:
    <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>
    """
    def get_organisations(self):
        """
        This method creates organisation objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of organisations
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['Organisation']
        
        # Loop over rows of excel sheet
        first_row = True
        organisations = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                description = row[1].value

                if type(row[2].value) == str:
                    pages = [page.strip() for page in row[2].value.split(";")]
                else:
                    pages = []

                location_title = row[3].value
                location_description = row[4].value

                # Create organisation object and add to organisation dictionary
                organisation = VPOrganisation.VPOrganisation(Config.CATALOG_URL, title, description, location_title, location_description, pages)
                organisations[organisation.TITLE] = organisation

        return organisations

    def get_biobanks(self):
        """
        This method creates biobank objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of biobanks
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['BiobankPatientRegistry']
        
        # Loop over rows of excel sheet
        first_row = True
        biobanks = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                description = row[1].value
                populationcoverage = row[2].value

                if type(row[3].value) == str:
                    themes = [theme.strip() for theme in row[3].value.split(";")]
                else:
                    themes = []

                publisher_name = row[4].value

                if type(row[5].value) == str:
                    pages = [page.strip() for page in row[5].value.split(";")]
                else:
                    pages = []

                resource_type = row[6].value

                # Create biobank object and add to biobank dictionary if it is a biobank
                if resource_type == "Biobank":
                    biobank = VPBiobank.VPBiobank(Config.CATALOG_URL, None, title, description, populationcoverage, themes, publisher_name, pages)
                    biobanks[biobank.TITLE] = biobank

        return biobanks

    def get_patientregistries(self):
        """
        This method creates patient registry objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of patientregistries
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['BiobankPatientRegistry']
        
        # Loop over rows of excel sheet
        first_row = True
        patientregistries = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                description = row[1].value
                populationcoverage = row[2].value

                if type(row[3].value) == str:
                    themes = [theme.strip() for theme in row[3].value.split(";")]
                else:
                    themes = []

                publisher_name = row[4].value

                if type(row[5].value) == str:
                    pages = [page.strip() for page in row[5].value.split(";")]
                else:
                    pages = []

                resource_type = row[6].value

                # Create patient registry object and add to patientregistry dictionary if it is a patientregistry
                if resource_type == "Patient registry":
                    patientregistry = VPPatientregistry.VPPatientregistry(Config.CATALOG_URL, None, title, description, populationcoverage, themes, publisher_name, pages)
                    patientregistries[patientregistry.TITLE] = patientregistry

        return patientregistries

    def get_datasets(self):
        """
        This method creates dataset objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of datasets
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['Dataset']
        
        # Loop over rows of excel sheet
        first_row = True
        datasets = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                description = row[1].value

                if type(row[2].value) == str:
                    themes = [theme.strip() for theme in row[2].value.split(";")]
                else:
                    themes = []

                vpconnection = row[3].value
                license = row[4].value
                
                if type(row[5].value) == str:
                    related = [item.strip() for item in row[5].value.split(";")]
                else:
                    related = []

                version = row[6].value
                if version == None:
                    version = "1"

                if type(row[7].value) == str:
                    keywords = [item.strip() for item in row[7].value.split(";")]
                else:
                    keywords = []

                publisher_name = row[8].value
                page = row[9].value

                # Create dataset object and add to dataset dictionary
                dataset = VPDataset.VPDataset(Config.CATALOG_URL, title, description, keywords, themes, 
                                              None, publisher_name, "en", license, page, None, 
                                              vpconnection, related, version)
                datasets[dataset.TITLE] = dataset

        return datasets

    def get_distributions(self):
        """
        This method creates distribution objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of distributions
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['Distribution']
        
        # Loop over rows of excel sheet
        first_row = True
        distributions = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                dataset_title = row[1].value
                description = row[2].value
                url = row[3].value
                url_type = row[4].value
                license = row[5].value
                version = row[6].value
                mediatype = row[7].value
                publisher_name = row[8].value
                if type(row[9].value) == str:
                    ispartof = [item.strip() for item in row[9].value.split(";")]
                else:
                    ispartof = []

                # Create dataset object and add to dataset dictionary
                distribution = VPDistribution.VPDistribution(None, title, dataset_title, description,
                                                             None, publisher_name, license, version, url, url_type,
                                                             mediatype, ispartof)
                distributions[distribution.TITLE] = distribution

        return distributions